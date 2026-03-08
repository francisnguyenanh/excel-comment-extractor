import json
import os
import re
import shutil
import time
import uuid
import zipfile
import xml.etree.ElementTree as ET
from functools import wraps

from flask import (Flask, after_this_request, render_template,
                   request, send_file, abort, session, redirect, url_for)
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'dev-secret-key-change-in-production')

UPLOAD_BASE = os.path.join(os.path.dirname(__file__), 'uploads')
os.makedirs(UPLOAD_BASE, exist_ok=True)

CONFIG_FILE = os.path.join(os.path.dirname(__file__), 'config.json')
FOLDER_TTL  = 86400  # seconds — folders older than 24 h will be deleted

SESSION_LIFETIME    = 1800   # seconds — auto logout after 30 min of inactivity
LOGIN_MAX_ATTEMPTS  = 5      # max failed attempts before blocking an IP
LOGIN_BLOCK_TTL     = 86400  # seconds — IP is blocked for 24 h

ALLOWED_EXTENSIONS = {'xlsx'}

# In-memory store for brute-force tracking: {ip: {'count': int, 'blocked_until': float}}
_login_attempts: dict = {}


def allowed_file(filename: str) -> bool:
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


# ---------------------------------------------------------------------------
# Config / auth helpers
# ---------------------------------------------------------------------------
def _load_config() -> dict:
    """Read config.json. Auto-create with default password 'admin123' if missing."""
    if not os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
            json.dump({'password': 'admin123'}, f, indent=2)
        return {'password': 'admin123'}
    with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
        return json.load(f)


def _check_password(plain: str) -> bool:
    config = _load_config()
    return plain == config.get('password', '')


# ---------------------------------------------------------------------------
# Brute-force protection helpers
# ---------------------------------------------------------------------------
def _get_client_ip() -> str:
    """Return the real client IP, respecting X-Forwarded-For if present."""
    return request.headers.get('X-Forwarded-For', request.remote_addr or '0.0.0.0').split(',')[0].strip()


def _is_ip_blocked(ip: str) -> tuple[bool, float]:
    """Return (blocked, seconds_remaining). Also clears expired entries."""
    entry = _login_attempts.get(ip)
    if not entry:
        return False, 0.0
    blocked_until = entry.get('blocked_until', 0)
    if blocked_until > time.time():
        return True, blocked_until - time.time()
    # Block expired — clean up
    _login_attempts.pop(ip, None)
    return False, 0.0


def _record_login_failure(ip: str) -> int:
    """Increment failure counter for IP. Block IP if limit reached. Returns current count."""
    entry = _login_attempts.setdefault(ip, {'count': 0, 'blocked_until': 0})
    entry['count'] += 1
    if entry['count'] >= LOGIN_MAX_ATTEMPTS:
        entry['blocked_until'] = time.time() + LOGIN_BLOCK_TTL
    return entry['count']


def _clear_login_attempts(ip: str) -> None:
    """Remove all failure records for IP after a successful login."""
    _login_attempts.pop(ip, None)


def _user_folder() -> str:
    """
    Return the upload folder path for the current session (named by user_folder_id).
    Creates the folder if it doesn't exist yet.
    """
    fid = session.get('user_folder_id')
    if not fid:
        fid = str(uuid.uuid4())
        session['user_folder_id'] = fid
    path = os.path.join(UPLOAD_BASE, fid)
    os.makedirs(path, exist_ok=True)
    return path


# ---------------------------------------------------------------------------
# Login required decorator
# ---------------------------------------------------------------------------
def login_required(f):
    @wraps(f)
    def _wrapped(*args, **kwargs):
        if not session.get('logged_in'):
            return redirect(url_for('login_page'))
        return f(*args, **kwargs)
    return _wrapped


# ---------------------------------------------------------------------------
# Idle session timeout
# ---------------------------------------------------------------------------
@app.before_request
def _check_session_idle():
    """Auto-logout when the session has been idle longer than SESSION_LIFETIME."""
    if request.endpoint in ('login_page', 'logout', 'static'):
        return
    if session.get('logged_in'):
        last = session.get('last_activity', 0)
        if time.time() - last > SESSION_LIFETIME:
            try:
                shutil.rmtree(_user_folder(), ignore_errors=True)
            except Exception:
                pass
            session.clear()
            return redirect(url_for('login_page'))
        session['last_activity'] = time.time()


# ---------------------------------------------------------------------------
# XML namespace constants
# ---------------------------------------------------------------------------
NS_WB  = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
NS_R   = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
NS_REL = 'http://schemas.openxmlformats.org/package/2006/relationships'
NS_TC  = 'http://schemas.microsoft.com/office/spreadsheetml/2018/threadedcomments'


# ---------------------------------------------------------------------------
# Step 1 – Map sheet path → (sheet_name, sheet_index)
# ---------------------------------------------------------------------------
def get_sheet_name_map(z: zipfile.ZipFile) -> dict:
    """
    Returns:
        { 'xl/worksheets/sheet3.xml': ('ER_0209_入庫対象リスト', 2) }
    Index is 0-based, preserving workbook order.
    """
    # Parse xl/workbook.xml
    with z.open('xl/workbook.xml') as f:
        wb_tree = ET.parse(f)
    sheets = wb_tree.findall(f'.//{{{NS_WB}}}sheet')

    # Parse xl/_rels/workbook.xml.rels
    with z.open('xl/_rels/workbook.xml.rels') as f:
        rels_tree = ET.parse(f)
    rels = {
        r.get('Id'): r.get('Target')
        for r in rels_tree.findall(f'.//{{{NS_REL}}}Relationship')
    }

    result = {}
    for idx, sheet in enumerate(sheets):
        rid = sheet.get(f'{{{NS_R}}}id')
        target = rels.get(rid, '')          # e.g. "worksheets/sheet3.xml"
        full_path = f'xl/{target}'
        result[full_path] = (sheet.get('name'), idx)

    return result


# ---------------------------------------------------------------------------
# Step 2 – Find comment XML files for a given sheet (threaded + legacy)
# ---------------------------------------------------------------------------
def get_comment_files(z: zipfile.ZipFile, sheet_path: str) -> dict:
    """
    Returns paths to both comment XML files for the given sheet:
        {
            'threaded': 'xl/threadedComments/threadedComment1.xml' or None,
            'legacy':   'xl/comments1.xml' or None,
        }
    'threaded'  → Excel native threaded comments (has done attribute)
    'legacy'    → classic comments / Google Sheets export (full text in <t>)
    """
    sheet_filename = sheet_path.split('/')[-1]
    rels_path = f'xl/worksheets/_rels/{sheet_filename}.rels'

    result = {'threaded': None, 'legacy': None}

    if rels_path not in z.namelist():
        return result

    with z.open(rels_path) as f:
        tree = ET.parse(f)

    for rel in tree.findall(f'.//{{{NS_REL}}}Relationship'):
        rel_type = rel.get('Type', '')
        target   = 'xl/' + rel.get('Target', '').replace('../', '')
        if 'threadedComment' in rel_type:
            result['threaded'] = target
        elif rel_type.endswith('/relationships/comments'):
            result['legacy'] = target

    return result


# ---------------------------------------------------------------------------
# Helper functions
# ---------------------------------------------------------------------------
def _extract_legacy_text(text_el) -> str:
    """
    Extract plain text from a legacy <text> element.
    Google Sheets export:  <text><t xml:space="preserve">...</t></text>
    Classic rich-text:     <text><r><rPr/><t>...</t></r></text>
    """
    if text_el is None:
        return ''
    # Direct <t> (Google Sheets / simple format)
    t_el = text_el.find(f'{{{NS_WB}}}t')
    if t_el is not None and t_el.text:
        return t_el.text
    # Rich-text <r><t> runs
    parts = []
    for r_el in text_el.findall(f'{{{NS_WB}}}r'):
        t_el2 = r_el.find(f'{{{NS_WB}}}t')
        if t_el2 is not None and t_el2.text:
            parts.append(t_el2.text)
    return ''.join(parts)


def _get_cell_value(ws, cell_ref: str):
    """Safe cell value lookup via openpyxl worksheet."""
    if ws is None or not cell_ref:
        return None
    try:
        val = ws[cell_ref].value
        return str(val) if val is not None else None
    except Exception:
        return None


def _build_comment_dict(sheet_name, sheet_index, cell_ref, cell_value,
                        raw_text, replies, is_resolved, note_type) -> dict:
    return {
        'sheet_name':     sheet_name,
        'cell_ref':       cell_ref,
        'cell_value':     cell_value,
        'comment_text':   raw_text,
        'replies':        replies,
        'first_datetime': replies[0]['datetime'] if replies else '',
        'is_resolved':    is_resolved,
        'sheet_index':    sheet_index,
        'note_type':      note_type,   # 'threaded' | 'note'
    }


# ---------------------------------------------------------------------------
# Step 3 – Parse raw comment text into individual replies
# ---------------------------------------------------------------------------
def parse_replies(raw_text: str) -> list:
    """
    Parse a threaded-comment text blob into a list of reply dicts:
        {comment_id, author, datetime, content}
    """
    # Normalize Windows CRLF → LF (XML inside .xlsx uses \r\n)
    raw_text = raw_text.replace('\r\n', '\n').replace('\r', '\n')

    # Strip Excel compatibility header that some clients prepend
    if '[Threaded comment]' in raw_text:
        idx = raw_text.find('======')
        if idx >= 0:
            raw_text = raw_text[idx:]

    # Remove leading "======" separator
    raw_text = raw_text.lstrip('=').strip()

    # Split on reply separators (4+ dashes on their own line)
    segments = re.split(r'\n[-]{4,}\n', raw_text)

    replies = []
    for segment in segments:
        lines = segment.strip().splitlines()
        if not lines:
            continue

        comment_id = ''
        author = ''
        dt_str = ''
        i = 0

        # Line 1: ID#...
        if lines[i].startswith('ID#'):
            comment_id = lines[i].strip()
            i += 1

        # Line 2: "Author Name    (YYYY-MM-DD HH:MM:SS)"
        if i < len(lines):
            m = re.match(
                r'^(.+?)\s{2,}\((\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2})\)\s*$',
                lines[i]
            )
            if m:
                author = m.group(1).strip()
                dt_str = m.group(2)
                i += 1

        content = '\n'.join(lines[i:]).strip()

        replies.append({
            'comment_id': comment_id,
            'author': author,
            'datetime': dt_str,
            'content': content,
        })

    return replies


# ---------------------------------------------------------------------------
# Main extraction function
# ---------------------------------------------------------------------------
def extract_comments(filepath: str) -> list:
    """
    Open an .xlsx file and return all comments (threaded + legacy/Google Sheets)
    as a list of dicts including a 'note_type' field ('threaded' | 'note').
    """
    wb = load_workbook(filepath, data_only=True)
    results = []

    with zipfile.ZipFile(filepath, 'r') as z:
        sheet_map = get_sheet_name_map(z)

        for sheet_path, (sheet_name, sheet_index) in sheet_map.items():
            comment_files = get_comment_files(z, sheet_path)
            ws = wb[sheet_name] if sheet_name in wb.sheetnames else None

            # ── 1. Threaded Comments (Excel native) ──────────────────────
            tc_path = comment_files['threaded']
            has_threaded = False
            if tc_path and tc_path in z.namelist():
                with z.open(tc_path) as f:
                    tc_tree = ET.parse(f)

                tc_elements = tc_tree.findall(f'.//{{{NS_TC}}}threadedComment')
                if not tc_elements:
                    tc_elements = tc_tree.findall('.//threadedComment')

                for tc_el in tc_elements:
                    cell_ref    = tc_el.get('ref', '')
                    is_resolved = tc_el.get('done', '0') == '1'

                    text_el = tc_el.find(f'{{{NS_TC}}}text')
                    if text_el is None:
                        text_el = tc_el.find('text')
                    raw_text = (text_el.text or '') if text_el is not None else ''

                    replies = parse_replies(raw_text)
                    if not replies:
                        continue

                    results.append(_build_comment_dict(
                        sheet_name, sheet_index, cell_ref,
                        _get_cell_value(ws, cell_ref),
                        raw_text, replies, is_resolved, 'threaded',
                    ))
                    has_threaded = True

            # ── 2. Legacy / Google Sheets Comments ───────────────────────
            legacy_path = comment_files['legacy']
            if legacy_path and legacy_path in z.namelist():
                with z.open(legacy_path) as f:
                    legacy_tree = ET.parse(f)

                authors_el  = legacy_tree.find(f'.//{{{NS_WB}}}authors')
                author_list = []
                if authors_el is not None:
                    for a_el in authors_el.findall(f'{{{NS_WB}}}author'):
                        author_list.append(a_el.text or '')

                comment_list_el = legacy_tree.find(f'.//{{{NS_WB}}}commentList')
                if comment_list_el is None:
                    continue

                for comment_el in comment_list_el.findall(f'{{{NS_WB}}}comment'):
                    cell_ref    = comment_el.get('ref', '')
                    author_id   = int(comment_el.get('authorId', '0'))
                    author_name = author_list[author_id] if author_id < len(author_list) else ''

                    # Skip threaded comment wrapper entries (Excel native files)
                    if author_name.startswith('tc='):
                        continue

                    text_el  = comment_el.find(f'{{{NS_WB}}}text')
                    raw_text = _extract_legacy_text(text_el)

                    # Skip wrappers that point to threaded comments
                    if '[Threaded comment]' in raw_text or not raw_text.strip():
                        continue

                    replies = parse_replies(raw_text)
                    if not replies:
                        continue

                    results.append(_build_comment_dict(
                        sheet_name, sheet_index, cell_ref,
                        _get_cell_value(ws, cell_ref),
                        raw_text, replies, False, 'note',
                    ))

    results.sort(key=lambda x: (x['sheet_index'], x['first_datetime'] or ''))
    return results


# ---------------------------------------------------------------------------
# Excel export builder
# ---------------------------------------------------------------------------
def build_excel(comments: list, output_path: str) -> None:
    """Generate a formatted .xlsx file from extracted comments."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Comments'

    header_font   = Font(bold=True, color='FFFFFF', size=10)
    header_fill   = PatternFill('solid', fgColor='1A56DB')
    resolved_fill = PatternFill('solid', fgColor='D1FAE5')
    open_fill     = PatternFill('solid', fgColor='DBEAFE')
    alt_fill      = PatternFill('solid', fgColor='F8FAFF')
    center        = Alignment(horizontal='center', vertical='top', wrap_text=True)
    wrap_top      = Alignment(vertical='top', wrap_text=True)
    thin_side     = Side(style='thin', color='D1D5DB')
    thin_border   = Border(left=thin_side, right=thin_side,
                           top=thin_side, bottom=thin_side)

    headers = [
        '#', 'Sheet', 'Ô', 'Nội dung ô', 'Trạng thái',
        'Reply #', 'Comment ID', 'Tên tác giả', 'Thời gian', 'Nội dung reply',
    ]
    col_widths = [5, 28, 7, 28, 12, 8, 22, 26, 20, 60]

    ws.append(headers)
    for col_idx, (hdr, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = width

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = 'A2'

    data_row = 2
    for i, c in enumerate(comments, 1):
        status_label = 'Resolved' if c['is_resolved'] else 'Open'
        row_fill     = resolved_fill if c['is_resolved'] else open_fill

        for j, reply in enumerate(c['replies'], 1):
            ws.append([
                i,
                c['sheet_name'],
                c['cell_ref'],
                c['cell_value'] or '',
                status_label,
                j,
                reply['comment_id'],
                reply['author'],
                reply['datetime'],
                reply['content'],
            ])
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=data_row, column=col_idx)
                cell.border = thin_border
                cell.alignment = center if col_idx in (1, 3, 5, 6) else wrap_top
                if col_idx == 5:
                    cell.fill = row_fill
                elif data_row % 2 == 0:
                    cell.fill = alt_fill
            data_row += 1

    ws.auto_filter.ref = f'A1:{ws.cell(row=1, column=len(headers)).coordinate}'
    wb.save(output_path)


# ---------------------------------------------------------------------------
# Cleanup helpers
# ---------------------------------------------------------------------------
def _cleanup_expired_folders():
    """
    Scan UPLOAD_BASE and delete sub-folders whose mtime is older than FOLDER_TTL.
    Called once per GET / request to garbage-collect abandoned sessions.
    """
    if not os.path.exists(UPLOAD_BASE):
        return
    cutoff = time.time() - FOLDER_TTL
    try:
        for entry in os.scandir(UPLOAD_BASE):
            if entry.is_dir() and entry.stat().st_mtime < cutoff:
                shutil.rmtree(entry.path, ignore_errors=True)
    except Exception:
        pass


def _cleanup_user_folder(keep_token: str | None = None):
    """
    Delete all files in the current user's upload folder,
    except comments_{keep_token}.xlsx (still needed for download).
    """
    folder = _user_folder()
    keep   = f'comments_{keep_token}.xlsx' if keep_token else None
    try:
        for name in os.listdir(folder):
            if name == keep:
                continue
            fp = os.path.join(folder, name)
            if os.path.isfile(fp):
                os.remove(fp)
    except Exception:
        pass


# ---------------------------------------------------------------------------
# Flask routes
# ---------------------------------------------------------------------------
@app.route('/login', methods=['GET', 'POST'])
def login_page():
    if session.get('logged_in'):
        return redirect(url_for('index'))

    ip = _get_client_ip()
    blocked, remaining_secs = _is_ip_blocked(ip)
    error = None

    if blocked:
        hours   = int(remaining_secs) // 3600
        minutes = (int(remaining_secs) % 3600) // 60
        error   = (f'Thiết bị của bạn đã bị khoá do nhập sai mật khẩu quá {LOGIN_MAX_ATTEMPTS} lần. '
                   f'Vui lòng thử lại sau {hours} giờ {minutes} phút.')
        return render_template('login.html', error=error, blocked=True)

    if request.method == 'POST':
        if _check_password(request.form.get('password', '')):
            _clear_login_attempts(ip)
            session.clear()
            session['logged_in']      = True
            session['user_folder_id'] = str(uuid.uuid4())
            session['last_activity']  = time.time()
            return redirect(url_for('index'))
        # Wrong password
        count = _record_login_failure(ip)
        blocked_now, _ = _is_ip_blocked(ip)
        if blocked_now:
            error = (f'Bạn đã nhập sai {LOGIN_MAX_ATTEMPTS} lần. '
                     f'Thiết bị này bị khoá trong 24 giờ.')
        else:
            remaining = LOGIN_MAX_ATTEMPTS - count
            error = f'Mật khẩu không đúng. Còn {remaining} lần thử trước khi bị khoá.'

    return render_template('login.html', error=error, blocked=blocked)


@app.route('/logout')
def logout():
    try:
        shutil.rmtree(_user_folder(), ignore_errors=True)
    except Exception:
        pass
    session.clear()
    return redirect(url_for('login_page'))


@app.route('/')
@login_required
def index():
    _cleanup_expired_folders()

    result_token = session.pop('result_token', None)
    comments, error, download_token = [], None, None

    if result_token:
        folder      = _user_folder()
        result_path = os.path.join(folder, f'result_{result_token}.json')
        if os.path.exists(result_path):
            with open(result_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            comments       = data.get('comments', [])
            error          = data.get('error')
            download_token = data.get('download_token')
            os.remove(result_path)

        _cleanup_user_folder(keep_token=download_token)

        return render_template('index.html', comments=comments,
                               error=error, download_token=download_token)

    _cleanup_user_folder(keep_token=None)
    return render_template('index.html')


@app.route('/upload', methods=['POST'])
@login_required
def upload():
    folder = _user_folder()

    def _redirect_with(error=None, comments=None, download_token=None):
        token       = str(uuid.uuid4())
        result_path = os.path.join(folder, f'result_{token}.json')
        with open(result_path, 'w', encoding='utf-8') as f:
            json.dump({
                'comments':       comments or [],
                'error':          error,
                'download_token': download_token,
            }, f, ensure_ascii=False)
        session['result_token'] = token
        return redirect(url_for('index'))

    if 'file' not in request.files:
        return _redirect_with(error='Không tìm thấy file trong request.')

    file = request.files['file']

    if file.filename == '':
        return _redirect_with(error='Chưa chọn file.')

    if not allowed_file(file.filename):
        return _redirect_with(error='Chỉ chấp nhận file .xlsx.')

    token       = str(uuid.uuid4())
    upload_path = os.path.join(folder, f'upload_{token}.xlsx')
    export_path = os.path.join(folder, f'comments_{token}.xlsx')

    error = None
    comments = []
    download_token = None

    try:
        file.save(upload_path)
        comments = extract_comments(upload_path)
        if comments:
            build_excel(comments, export_path)
            download_token = token
    except Exception as exc:
        error = f'Lỗi khi xử lý file: {exc}'
    finally:
        if os.path.exists(upload_path):
            os.remove(upload_path)

    return _redirect_with(error=error, comments=comments, download_token=download_token)


@app.route('/download/<token>')
@login_required
def download(token):
    if not re.match(
        r'^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}$',
        token,
    ):
        abort(400)

    filepath = os.path.join(_user_folder(), f'comments_{token}.xlsx')
    if not os.path.exists(filepath):
        abort(404)

    @after_this_request
    def remove_export(response):
        try:
            os.remove(filepath)
        except Exception:
            pass
        return response

    return send_file(
        filepath,
        as_attachment=True,
        download_name='comments_extracted.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


if __name__ == '__main__':
    app.run(debug=True)
